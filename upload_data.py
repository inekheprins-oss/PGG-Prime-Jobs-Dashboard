"""
PGG Dashboard — Data Uploader
==============================
Double-click this file to read your Excel and push fresh data to GitHub.
The dashboard URL updates within ~60 seconds for all users.

SETUP (one time only):
  1. Fill in GITHUB_USERNAME, GITHUB_TOKEN and EXCEL_PATH below
  2. Save the file
  3. From then on, just double-click to refresh

HOW TO GET A GITHUB TOKEN:
  1. Log in to github.com
  2. Click your profile picture (top right) → Settings
  3. Scroll to the bottom → Developer settings
  4. Personal access tokens → Tokens (classic) → Generate new token (classic)
  5. Give it a name like "dashboard-upload"
  6. Tick the "repo" checkbox (full repo access)
  7. Click Generate token — COPY IT NOW (you only see it once)
  8. Paste it below as GITHUB_TOKEN
"""

import json, base64, urllib.request, urllib.error, sys, os
from datetime import datetime

# ============================================================
# CONFIGURE THESE THREE LINES (one-time setup)
# ============================================================
GITHUB_USERNAME = "YOUR_GITHUB_USERNAME"       # e.g. "pgg-dashboard"
GITHUB_TOKEN    = "YOUR_GITHUB_TOKEN"           # paste token here
EXCEL_PATH      = r"C:\path\to\Prime_Jobs_per_Day.xlsx"  # full path to your Excel
# ============================================================

REPO_NAME  = "pgg-dashboard"
DATA_FILE  = "data.json"
BRANCH     = "main"

def log(msg):
    print(msg, flush=True)

def read_excel(path):
    """Read Monthly Target and Trading Days sheets from the Excel file."""
    try:
        import openpyxl
    except ImportError:
        log("Installing openpyxl...")
        import subprocess
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
        import openpyxl

    log(f"Reading: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # ---- Monthly Target sheet ----
    ws = wb["Monthly Target"]
    centres = []
    current_region = ""
    rows = list(ws.iter_rows(values_only=True))

    for i, row in enumerate(rows):
        if i == 0:
            continue  # skip header
        if not row or all(v is None for v in row):
            continue
        plant = row[0]
        name  = row[1]
        if isinstance(plant, str) and plant.strip() and not name:
            current_region = plant.strip()
            continue
        if isinstance(plant, int) and name:
            centres.append({
                "plant":              plant,
                "name":               str(name).strip(),
                "regional":           str(row[2]).strip() if row[2] else "",
                "region":             current_region,
                "primeTarget":        float(row[4]  or 0),
                "gap":                float(row[5]  or 0),
                "dailyTarget":        round(float(row[6]  or 0), 2),
                "primeCompletedMTD":  float(row[7]  or 0),
                "primeInvoicedMTD":   float(row[8]  or 0),
                "filmCompletedMTD":   float(row[9]  or 0),
                "filmInvoicedMTD":    float(row[10] or 0),
                "adasCompletedMTD":   float(row[11] or 0),
                "adasInvoicedMTD":    float(row[12] or 0),
                "otherCompletedMTD":  float(row[13] or 0),
                "otherInvoicedMTD":   float(row[14] or 0),
                "dailyPrimeGap":      round(float(row[15] or 0), 2),
                "primeInvoicedToday": float(row[17] or 0),
            })

    if not centres:
        raise ValueError("No fitment centres found in 'Monthly Target' sheet. Check the file.")

    # ---- Trading Days sheet ----
    trading_info = {"monthDaysLeft": 0, "tradingDayOfMonth": 0, "totalTradingDaysInMonth": 20}
    try:
        ws2 = wb["Trading Days"]
        rows2 = list(ws2.iter_rows(values_only=True, max_row=3))
        if len(rows2) > 1 and rows2[1][1] is not None:
            trading_info["monthDaysLeft"] = int(rows2[1][1])
        if len(rows2) > 1 and rows2[1][2] is not None:
            trading_info["tradingDayOfMonth"] = int(rows2[1][2]) if rows2[1][2] else 0
        # Try to figure out trading day of month from the calendar rows
        # Column F (index 5) = Day name, Column G (index 6) = Trading Day (Month)
        today_str = datetime.now().strftime("%Y-%m-%d")
        for row in rows2[1:]:
            if row[4] is not None:
                try:
                    from datetime import date
                    d = row[4]
                    if hasattr(d, 'date'):
                        d = d.date()
                    if str(d) == today_str and row[6] is not None:
                        trading_info["tradingDayOfMonth"] = int(row[6])
                        break
                except:
                    pass
    except Exception as e:
        log(f"  (Trading Days sheet: {e} — using defaults)")

    log(f"  Found {len(centres)} fitment centres")
    log(f"  Trading days left this month: {trading_info['monthDaysLeft']}")
    return centres, trading_info


def get_file_sha(token, username, repo, filepath, branch):
    """Get the current SHA of a file (needed to update it on GitHub)."""
    url = f"https://api.github.com/repos/{username}/{repo}/contents/{filepath}?ref={branch}"
    req = urllib.request.Request(url)
    req.add_header("Authorization", f"token {token}")
    req.add_header("Accept", "application/vnd.github.v3+json")
    try:
        with urllib.request.urlopen(req) as resp:
            data = json.loads(resp.read())
            return data.get("sha")
    except urllib.error.HTTPError as e:
        if e.code == 404:
            return None  # file doesn't exist yet — that's fine
        raise


def upload_file(token, username, repo, filepath, content_str, branch, sha=None):
    """Create or update a file on GitHub."""
    url = f"https://api.github.com/repos/{username}/{repo}/contents/{filepath}"
    content_b64 = base64.b64encode(content_str.encode("utf-8")).decode("utf-8")
    payload = {
        "message": f"Data refresh {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "content": content_b64,
        "branch":  branch,
    }
    if sha:
        payload["sha"] = sha

    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, data=data, method="PUT")
    req.add_header("Authorization", f"token {token}")
    req.add_header("Content-Type", "application/json")
    req.add_header("Accept", "application/vnd.github.v3+json")

    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read())


def main():
    log("=" * 50)
    log("  PGG Dashboard — Data Uploader")
    log("=" * 50)

    # Check config
    if "YOUR_GITHUB" in GITHUB_USERNAME or "YOUR_GITHUB" in GITHUB_TOKEN:
        log("\n⚠  ERROR: You haven't filled in your GitHub details yet.")
        log("   Open this file in Notepad and set:")
        log("   - GITHUB_USERNAME")
        log("   - GITHUB_TOKEN")
        log("   - EXCEL_PATH")
        input("\nPress Enter to close...")
        return

    if not os.path.exists(EXCEL_PATH):
        log(f"\n⚠  ERROR: Excel file not found at:\n   {EXCEL_PATH}")
        log("   Update EXCEL_PATH in this script to the correct location.")
        input("\nPress Enter to close...")
        return

    try:
        # 1. Read Excel
        log("\n[1/3] Reading Excel file...")
        centres, trading_info = read_excel(EXCEL_PATH)

        # 2. Build data.json
        log("\n[2/3] Building data package...")
        data = {
            "lastUpdated": datetime.now().isoformat(timespec="seconds"),
            "tradingInfo": trading_info,
            "centres": centres
        }
        data_str = json.dumps(data, indent=2)
        log(f"  Package size: {len(data_str):,} bytes")

        # 3. Upload to GitHub
        log("\n[3/3] Uploading to GitHub...")
        sha = get_file_sha(GITHUB_TOKEN, GITHUB_USERNAME, REPO_NAME, DATA_FILE, BRANCH)
        action = "Updating" if sha else "Creating"
        log(f"  {action} {DATA_FILE}...")
        upload_file(GITHUB_TOKEN, GITHUB_USERNAME, REPO_NAME, DATA_FILE, data_str, BRANCH, sha)

        log("\n✓ SUCCESS!")
        log(f"  Dashboard will reflect new data within ~60 seconds.")
        log(f"  URL: https://{GITHUB_USERNAME}.github.io/{REPO_NAME}/")
        log(f"  Updated at: {datetime.now().strftime('%H:%M:%S')}")

    except Exception as e:
        log(f"\n✕ ERROR: {e}")
        import traceback
        traceback.print_exc()

    input("\nPress Enter to close...")


if __name__ == "__main__":
    main()
